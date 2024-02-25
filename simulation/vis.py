from elements import *
import openpyxl as xl # openpyxl
from datetime import datetime
import os
import matplotlib.pyplot as plt # matplotlib
import numpy as np # numpy

### functions ###
def add_header(ws) :
  ws.append(["time", "people at queue", "people get on the bus", "waiting time"])

def con_excel() :
  wb = xl.Workbook()
  ws = wb.active
  ws.title = "0"
  add_header(ws)
  for i in range(1, len(Stop.list_location)-1) :
    ws = wb.create_sheet(str(i))
    add_header(ws)
  return wb

def des_excel(wb) :
  timestamp = datetime.now().strftime("%m%d-%H%M")
  try :
    os.mkdir("excel")
  except :
    pass
  wb.save(f"excel/{timestamp}.xlsx")
  wb.close()

def to_excel_ppl(wb, t) :  # add data to excel
  for stop in Stop.list_obj[:-1] :
    ws = wb[str(Stop.list_obj.index(stop))]
    row = [t, len(stop.user_list)]
    if (stop.current_on != 0) :
      row.append(stop.current_on)
    ws.append(row)
    stop.current_on = 0

def to_excel_waiting(wb) :
  for user in User.list_obj :
    ws = wb[str(user.stop)]
    ws.cell(row = user.enqueue_time + 2, column = 4).value = user.waiting_time
    # ws.cell(row = user.enqueue_time + 2, column = 4).value = user.waiting_num_bus
  for stop in Stop.list_obj[:-1] :  # remove waiting time of all the users that are still waiting at the queue
    ws = wb[str(Stop.list_obj.index(stop))]
    for user in stop.user_list :
      ws.cell(row = user.enqueue_time + 2, column = 4).value = None

def plt_waiting_time() :
  fig, axes = plt.subplots(3, 2)
  axes = axes.flatten()

  for stop in Stop.list_obj[:-1] :
    index = Stop.list_obj.index(stop)
    x = []
    y = []
    z = []
    ax = axes[index]
    for user in User.list_obj :
      if (user.stop != index) :
        continue # only search for the current stop
      if user not in stop.user_list :  # if not queueing
        x.append(user.enqueue_time)
        # z.append(user.waiting_num_bus)
        y.append(user.waiting_time)
        if (user.waiting_time != 0) :
          ax.axline((user.enqueue_time, user.waiting_time), slope=-1, lw=.5)
    ax.scatter(x, y, s=10)
    # ax.plot(x, z)
    ax.set_xlabel('time')
    ax.set_ylabel('waiting time')
    ax.set_title(index)
    ax.set_xlim(0, MAX_TIME+1)
    ax.set_ylim(0, max(BUS_CYCLE + 50, np.max(y)))
    t = 0
    for leave_time in stop.leave_time_list :  # to show when the bus leave
      ax.axvline(leave_time, color='red', linestyle='--')
      y = leave_time - t
      ax.plot((t, leave_time), (y, y), color='green', linestyle=':')
      t = leave_time

  plt.tight_layout()
  timestamp = datetime.now().strftime("%m%d-%H%M")
  plt.savefig(f'excel/{timestamp}.png')
  plt.show()

def is_fit(m, c, x, y) -> bool :  # see if the point(user) on the line of y=mx+c 
  return (y == m * x + c)

def next_x(x, arr) -> int : # return the smallest element that is larger than x
  for ele in arr :  # arr is sorted
    if ele > x :
      return ele
  return -1

def next_leave_time_diff(next_leave, leave_time) -> int :  # return the diff time of next leave time and the prev leave time
  index = leave_time.index(next_leave) -1
  if (index == -1) :
    return next_leave  # next_leave - 0
  else :
    return next_leave - leave_time[index]


##### not yet done
#
# def add_line(ax, x1, y1, waiting_time, enqueue_time, leave_time) -> None : # add a line for stops except the first stop
#   if (y1 == 0) :
#     m = 0
#   else :
#     m = -1
#   c = y1 - m * x1  # y=mx+c

#   x = x1
#   next_leave = next_x(x1, leave_time)
#   while (True) :
#     next_enqueue = next_x(x, enqueue_time)
#     if (next_enqueue == -1) :  # end at the last enqueue
#       ax.plot([x1, next_enqueue], [y1, m * next_enqueue + c], lw=.5)
#       return
    

#     elif (not is_fit(m, c, next_enqueue, next_waiting)) :
#         x = next_enqueue



#     elif (next_leave < next_enqueue) :  # inc when meets the next leave time
#       x2 = next_leave
#       y2 = m * x2 + c
#       ax.plot([x1, x2-1], [y1, y2], lw=.5)  # line stops just before the bus leave
#       y3 = next_leave_time_diff(next_leave, leave_time)
#       ax.plot([x2-1, x2], [y2, y3], lw=.5)  # line of the increase
#       add_line(ax, x2, y3, waiting_time, enqueue_time, leave_time)
#       return
    
#     else :  # a person enqueue before the bus arrive
#       next_waiting = waiting_time[enqueue_time.index(next_enqueue)]
#       if (is_fit(m, c, next_enqueue, next_waiting)) :
#         x = next_enqueue
#       else :  # inc when the next user gets on the other bus
#         y2 = waiting_time[enqueue_time.index(x)]
#         ax.plot([x1, x], [y1, y2], lw=.5)  # line stops at last person enqueue for that bus
#         y3 = next_leave_time_diff(next_leave, leave_time)
#         ax.plot([x, x+1], [y2, y3], lw=.5)  # line of the increase
#         add_line(ax, x+1, y3, waiting_time, enqueue_time, leave_time)
#         return
#
##### not yet done
